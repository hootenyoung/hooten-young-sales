"""Parser for the QuickBooks "Sales by Product/Service Detail" xlsx report.

Layout (hierarchical, with subtotal rows interspersed):

    Row 0:  Company name ("WI, Inc.")
    Row 1:  Report title
    Row 2:  Period label (e.g. "April 2026")
    Row 3:  blank
    Row 4:  Header row: '', Date, Transaction Type, Num, Customer,
            Memo/Description, Qty, Sales Price, Amount, Balance,
            P.O. Number
    Row 5:  Brand header (col 0 = "Hooten & Young")
    Row 6:  Product group header (col 0 = product name, others blank)
    Rows 7+: Invoice line rows (col 0 blank, col 1 = date, ...)
            interspersed with "Total for ..." subtotal rows
    Last:   "TOTAL" grand total row + generation timestamp

The parser ignores subtotals, headers, and grand totals — only invoice
line rows (those with a parseable date in col 1) are captured.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from hy_sales.parsers.canonical import ParsedSalesInvoice, ParsedSalesLine

# Column indices in the report. 0-based.
_COL_DATE = 1
_COL_TRANSACTION_TYPE = 2
_COL_NUM = 3
_COL_CUSTOMER = 4
_COL_MEMO = 5
_COL_QTY = 6
_COL_SALES_PRICE = 7
_COL_AMOUNT = 8
_COL_PO_NUMBER = 10


def parse_sales_xlsx(path: Path) -> list[ParsedSalesInvoice]:
    """Parse a QuickBooks 'Sales by Product/Service Detail' xlsx file.

    Returns invoices grouped by 'Num' (e.g. "SI-012682"). Lines are
    assigned 1-based ``line_seq`` values in file order — the same
    product can legitimately appear twice on one invoice.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []

    invoices: dict[str, ParsedSalesInvoice] = {}
    line_seq_per_invoice: dict[str, int] = {}

    for row in ws.iter_rows(values_only=True):
        if not _is_data_row(row):
            continue

        invoice_ref = str(row[_COL_NUM]).strip()
        invoice_date = _parse_date(row[_COL_DATE])
        if invoice_date is None or not invoice_ref:
            continue

        memo = _to_str(row[_COL_MEMO])
        if not memo:
            continue  # safety: no memo = no product, skip

        transaction_type = _normalize_transaction_type(row[_COL_TRANSACTION_TYPE])
        customer = _to_str(row[_COL_CUSTOMER]) or ""
        po_number = _to_str(row[_COL_PO_NUMBER])
        quantity = _to_decimal(row[_COL_QTY])
        sales_price = _to_decimal(row[_COL_SALES_PRICE])
        amount = _to_decimal(row[_COL_AMOUNT])

        if invoice_ref not in invoices:
            invoices[invoice_ref] = ParsedSalesInvoice(
                invoice_ref=invoice_ref,
                invoice_date=invoice_date,
                transaction_type=transaction_type,
                customer_raw_text=customer,
                po_number=po_number,
                lines=[],
            )
            line_seq_per_invoice[invoice_ref] = 0

        line_seq_per_invoice[invoice_ref] += 1
        invoices[invoice_ref].lines.append(
            ParsedSalesLine(
                line_seq=line_seq_per_invoice[invoice_ref],
                product_raw_text=memo,
                quantity=quantity,
                sales_price=sales_price,
                amount=amount,
            )
        )

    return list(invoices.values())


def _is_data_row(row: tuple[Any, ...]) -> bool:
    """Return True if this row looks like an invoice line (not a header/subtotal)."""
    if len(row) <= _COL_NUM:
        return False
    # Brand/product/subtotal rows have content in col 0 and a blank col 1.
    # Data rows have col 0 blank and a date in col 1.
    if row[0] is not None and str(row[0]).strip():
        return False
    # Date column must be present and parseable.
    if _parse_date(row[_COL_DATE]) is None:
        return False
    # Num must be present.
    if row[_COL_NUM] is None:
        return False
    num_str = str(row[_COL_NUM]).strip().lower()
    return num_str not in ("", "num")


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        # QuickBooks exports dates as 'MM/DD/YYYY'.
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return Decimal("0")
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return Decimal("0")
    return Decimal("0")


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _normalize_transaction_type(value: Any) -> str:
    """Map source-system transaction labels to our canonical values."""
    if value is None:
        return "invoice"
    label = str(value).strip().lower()
    if label == "invoice":
        return "invoice"
    if label in ("credit memo", "credit_memo"):
        return "credit_memo"
    return "other"
