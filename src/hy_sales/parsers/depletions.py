"""Parser for the depletions xlsx file (state x account x product x month pivot).

Supports four layouts that have appeared so far, auto-detected by the
position of the "Products" column in the dim header row:

* **iDIG Rolling Periods + Premises** (current; 10 dim columns):
  Dist States, Account Names, Address, City, State, Distributor code,
  Acct Counties, Acct Zips, OnOff Premises, Products — followed by N
  single ``9 Liter Equivs`` columns, one per month.

* **iDIG Rolling Periods** (previous; 9 dim columns):
  Dist States, Account Names, Address, City, State, Distributor code,
  Acct Counties, Acct Zips, Products — followed by N single
  ``9 Liter Equivs`` columns, one per month.

* **Older broker monthly** (8 dim columns):
  Dist States, Account Names, Address, City, State, Distributor code,
  Acct City State, Products — followed by paired
  ``(9 Liter Equivs, Physical Cases)`` columns per month.

* **MVP file** (6 dim columns):
  Dist States, Account Names, Address, City, Acct City State, Products —
  single ``9 Liter Equivs`` column per month.

Multi-month aggregate columns ("30 Months ...", "Diff", "Pct", etc.)
are ignored. Zero-valued cells ARE preserved — a stored zero records
"iDIG reported no depletions for that account/product/month", which is
distinct from "no data".
"""

from __future__ import annotations

import re
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from hy_sales.parsers.canonical import ParsedDepletionRow

_PERIOD_HEADER_RE = re.compile(r"^1\s*Month\s+(\d+)/(\d+)/(\d+)\s+thru", re.IGNORECASE)


def parse_depletions_xlsx(path: Path) -> list[ParsedDepletionRow]:
    """Parse a depletions xlsx file into long-format rows.

    Each returned row represents one (account, product, month) datapoint
    with cases_9l (and optionally cases_physical for layouts that pair
    the metrics). Zero-valued rows ARE returned — they record that the
    source explicitly reported zero depletions, not absence of data.
    """
    # NOTE: read_only=True would be faster, but iDIG exports declare a
    # stale "A1" worksheet dimension that makes read-only iter_rows yield
    # only the brand-title row. ~4k-row files don't need read-only mode.
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return []

    period_header_row = rows[1]
    dim_header_row = rows[2]

    dim_count = _detect_dim_count(dim_header_row)
    is_paired = _is_paired_metric_layout(dim_header_row, dim_count)
    value_columns = _identify_value_columns(period_header_row, dim_count, is_paired)
    layout = _resolve_layout(dim_count, dim_header_row)

    results: list[ParsedDepletionRow] = []

    for row in rows[3:]:
        if row is None or len(row) <= dim_count - 1:
            continue
        if _is_subtotal_row(row, dim_count):
            continue

        account_name = _to_str(row[1])
        product_raw = _to_str(row[dim_count - 1])
        if not account_name or not product_raw:
            continue

        dist_state_code = _normalize_state(row[layout["dist_state"]])
        state_code = (
            _normalize_state(row[layout["state"]])
            if layout["state"] is not None
            else dist_state_code
        )
        address = _to_str(row[layout["address"]])
        city = _to_str(row[layout["city"]])
        distributor_code = (
            _to_str(row[layout["distributor"]]) if layout["distributor"] is not None else None
        )
        county = _to_str(row[layout["county"]]) if layout["county"] is not None else None
        zip_code = _to_str(row[layout["zip"]]) if layout["zip"] is not None else None
        # premises_type is normalized by the ParsedDepletionRow validator
        # to 'ON' / 'OFF' / None — passing the raw cell value is fine.
        premises_type = _to_str(row[layout["premises"]]) if layout["premises"] is not None else None

        for col_idx, period_month in value_columns:
            if col_idx >= len(row):
                continue
            cases_9l = _to_decimal(row[col_idx])
            cases_physical: Decimal | None = None
            if is_paired and col_idx + 1 < len(row):
                cases_physical = _to_decimal(row[col_idx + 1])

            results.append(
                ParsedDepletionRow(
                    state_code=state_code,
                    dist_state_code=dist_state_code,
                    account_name=account_name,
                    account_address=address,
                    account_city=city,
                    account_county=county,
                    account_zip=zip_code,
                    distributor_code=distributor_code,
                    premises_type=premises_type,
                    product_raw_text=product_raw,
                    period_month=period_month,
                    cases_9l=cases_9l,
                    cases_physical=cases_physical,
                )
            )

    return results


def _detect_dim_count(dim_header_row: tuple[Any, ...]) -> int:
    """Return the number of dimension columns by locating the 'Products' column."""
    for i, cell in enumerate(dim_header_row):
        if cell is not None and str(cell).strip().lower() == "products":
            return i + 1
    # Fall back to the current iDIG layout (Rolling Periods + Premises).
    return 10


def _is_paired_metric_layout(dim_header_row: tuple[Any, ...], dim_count: int) -> bool:
    """Return True if metric columns are paired ``(9 Liter Equivs, Physical Cases)``."""
    if dim_count + 1 >= len(dim_header_row):
        return False
    first = dim_header_row[dim_count]
    second = dim_header_row[dim_count + 1]
    return (
        first is not None
        and "9 Liter" in str(first)
        and second is not None
        and "Physical" in str(second)
    )


def _resolve_layout(dim_count: int, dim_header_row: tuple[Any, ...]) -> dict[str, int | None]:
    """Map logical fields to column indices for the detected layout.

    Returns 0-based column indices (or None when the field is absent
    from this layout). ``dist_state`` is always col 0.
    """
    if dim_count == 10:
        # iDIG Rolling Periods + OnOff Premises (current).
        return {
            "dist_state": 0,
            "account_name": 1,
            "address": 2,
            "city": 3,
            "state": 4,
            "distributor": 5,
            "county": 6,
            "zip": 7,
            "premises": 8,
        }
    if dim_count == 9:
        # iDIG Rolling Periods (previous, pre-Premises).
        return {
            "dist_state": 0,
            "account_name": 1,
            "address": 2,
            "city": 3,
            "state": 4,
            "distributor": 5,
            "county": 6,
            "zip": 7,
            "premises": None,
        }
    if dim_count == 8:
        # Older broker monthly layout with paired metrics.
        return {
            "dist_state": 0,
            "account_name": 1,
            "address": 2,
            "city": 3,
            "state": 4,
            "distributor": 5,
            "county": None,
            "zip": None,
            "premises": None,
        }
    # 6-dim MVP layout — no separate State / Distributor / County / Zip.
    return {
        "dist_state": 0,
        "account_name": 1,
        "address": 2,
        "city": 3,
        "state": None,
        "distributor": None,
        "county": None,
        "zip": None,
        "premises": None,
    }


def _identify_value_columns(
    period_header_row: tuple[Any, ...],
    dim_count: int,
    is_paired: bool,
) -> list[tuple[int, date]]:
    """Return the list of monthly value-column starts as ``(col_idx, period_month)``.

    Stops at the first column whose header is not a 1-month period (e.g.
    "30 Months ...", "Diff", "Pct"), so aggregate columns are ignored.
    """
    stride = 2 if is_paired else 1
    out: list[tuple[int, date]] = []
    i = dim_count
    while i < len(period_header_row):
        header = period_header_row[i]
        if header is None:
            i += stride
            continue
        match = _PERIOD_HEADER_RE.match(str(header).strip())
        if not match:
            break  # hit the multi-month aggregate columns; stop.
        month = int(match.group(1))
        # group(2) is the day-of-month, which is always 1 in practice.
        year = int(match.group(3))
        try:
            period_month = date(year, month, 1)
        except ValueError:
            i += stride
            continue
        out.append((i, period_month))
        i += stride
    return out


def _is_subtotal_row(row: tuple[Any, ...], dim_count: int) -> bool:
    """Subtotal rows have 'Total' in the account-name or product columns."""
    account = row[1] if len(row) > 1 else None
    product = row[dim_count - 1] if len(row) > dim_count - 1 else None
    account_str = str(account).strip().lower() if account is not None else ""
    product_str = str(product).strip().lower() if product is not None else ""
    return account_str == "total" or product_str == "total"


def _normalize_state(value: Any) -> str | None:
    s = _to_str(value)
    if s is None:
        return None
    s = s.upper()
    return s[:2] if len(s) >= 2 else None


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned or cleaned == "--":
            return Decimal("0")
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return Decimal("0")
    return Decimal("0")
